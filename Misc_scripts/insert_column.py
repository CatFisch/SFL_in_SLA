#!/usr/bin/env python3

#Copyright 2021 Catharina Fischer

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

    #http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.


import glob
import os
import sys
import openpyxl
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import argparse
from copy import copy
import warnings
warnings.filterwarnings("ignore")

#create function for alignment of new column on column before
def merge_cells_in_new_column(ws, new_col_index, model_col_index):
    merged_cells = list(ws.merged_cells.ranges)
    for cr in merged_cells:
        if(cr.min_col==model_col_index):
            ws.merge_cells(start_row=cr.min_row,
                           start_column=new_col_index,
                           end_row=cr.max_row,
                           end_column=new_col_index)

#start function for inserting the column
def process_file(excel):
    wb = openpyxl.load_workbook(filename=excel, read_only=False)
    ws = wb.worksheets[0]
    base_name = excel.split('.')[0]
    col_inserted = False

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            if cell.value == 'repres': #add name of previous column here
                ws.insert_cols(cell.col_idx+1)
                merge_cells_in_new_column(ws, cell.col_idx+1, cell.col_idx)
                col_inserted = True
                ws.cell(row=1, column=cell.col_idx+1).value = 'process' #add name of new column here
                ws.cell(row=1, column=cell.col_idx + 1).font = copy(ws.cell(row=1, column=cell.col_idx).font)
                ws.cell(row=1, column=cell.col_idx + 1).fill = copy(ws.cell(row=1, column=cell.col_idx).fill)
                ws.cell(row=1, column=cell.col_idx + 1).border = copy(ws.cell(row=1, column=cell.col_idx).border)
                ws.cell(row=1, column=cell.col_idx + 1).alignment = copy(ws.cell(row=1, column=cell.col_idx).alignment)

    wb.save(excel)

    print(os.path.basename(base_name), 'has a new column now')


# end function for optionality of input files
parser = argparse.ArgumentParser()
parser.add_argument('--table', nargs='*')
args = parser.parse_args()

# insert column for a file in folder
if args.table:
    for f in args.table:
        process_file(f)
# insert column for all files in folder
else:
    excel_files = glob.glob(os.getcwd() + '/*.xlsx')


    for excel in excel_files:
        process_file(excel)

print('\n all selected files have an additional column now \n ฅ^•ﻌ•^ฅ')
