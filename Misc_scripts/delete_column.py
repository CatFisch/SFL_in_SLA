#!/usr/bin/env python3

#Copyright 2021 Catharina Fischer

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

    #http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.

import glob
import os
import openpyxl
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import warnings
warnings.filterwarnings("ignore")

excel_files = glob.glob(os.getcwd() + '/*.xlsx')

#delete "CONLL" sheet
for excel in excel_files:
    wb = openpyxl.load_workbook(filename=excel, read_only=False)
    sheetnames = wb.sheetnames
    CONLL = 'CONLL'
    if CONLL in sheetnames:
        del wb['CONLL']

    ws = wb.worksheets[0]

#call list with the name "DelCol.txt" that contains names of columns that should be deleted
    with open('DelCol.txt') as List:
        cols = List.read().splitlines()

#delete columns from list in respective excel files
    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            if cell.value not in cols:
                header_coordinate = coordinate_from_string(cell.coordinate)
                header_index = column_index_from_string(header_coordinate[0])
                ws.delete_cols(header_index)

    wb.save(excel)

