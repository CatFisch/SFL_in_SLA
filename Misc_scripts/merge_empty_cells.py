#!/usr/bin/env python3

#Copyright 2021 Catharina Fischer

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

    #http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.


import glob
import os
import openpyxl
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string

excel_files = glob.glob(os.getcwd() + '/*.xlsx')

#merge emplty cells with first one that contains content
for excel in excel_files:
    wb = openpyxl.load_workbook(filename=excel, read_only=False, data_only=True)
    ws = wb.worksheets[0]


    last_value = -1
    merging = False
    for row in ws.iter_rows(min_col=1, max_col=1): 
        cell = row[0]
        if not cell.value:
            if not merging:
                merging = True
                last_value = cell.row-1

        elif merging:
            ws.merge_cells(start_row=last_value,
                           start_column=1,
                           end_row=cell.row-1,
                           end_column=1)
            merging = False

    for row in ws.iter_rows(min_col=2, max_col=2): 
        cell = row[0]
        if not cell.value:
            if not merging:
                merging = True
                last_value = cell.row-1

        elif merging:
            ws.merge_cells(start_row=last_value,
                           start_column=2,
                           end_row=cell.row-1,
                           end_column=2)
            merging = False    

    for row in ws.iter_rows(min_col=3, max_col=3): 
        cell = row[0]
        if not cell.value:
            if not merging:
                merging = True
                last_value = cell.row-1

        elif merging:
            ws.merge_cells(start_row=last_value,
                           start_column=3,
                           end_row=cell.row-1,
                           end_column=3)
            merging = False    

    wb.save(excel)

