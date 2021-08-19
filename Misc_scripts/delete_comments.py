#!/usr/bin/env python3

#Copyright 2021 Catharina Fischer

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

    #http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.


import glob
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import warnings

warnings.filterwarnings("ignore")


excel_files = glob.glob(os.getcwd() + '/*.xlsx')


for excel in excel_files:
    wb = openpyxl.load_workbook(filename=excel, read_only=False)
    ws = wb.worksheets[0]


#delete all comments in sheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None
                
    wb.save(excel)

