#!/usr/bin/env python3

#Copyright 2021 Catharina Fischer

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

    #http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.


import glob
import os
import sys
import openpyxl
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import argparse
from copy import copy
import warnings
warnings.filterwarnings("ignore")

#create function for alignment of new column on "lemma"
def merge_cells_in_new_column(ws, new_col_index, model_col_index):
    merged_cells = list(ws.merged_cells.ranges)
    for cr in merged_cells:
        if(cr.min_col==model_col_index):
            ws.merge_cells(start_row=cr.min_row,
                           start_column=new_col_index,
                           end_row=cr.max_row,
                           end_column=new_col_index)
#start tagger function
def process_file(excel):
    wb = openpyxl.load_workbook(filename=excel, read_only=False)
    ws = wb.worksheets[0]
    base_name = excel.split('.')[0]
    col_inserted = False

    with open('ComVerbList.txt') as List:
        ComVerblist = List.read().splitlines()


    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            if cell.value == 'ZH1:ZH1lemma':
                lemma_col_index = coordinate_from_string(cell.coordinate)
            if cell.value == 'repres':
                repres_col_index = coordinate_from_string(cell.coordinate)
                repres = column_index_from_string(repres_col_index[0])
                ws.delete_cols(repres)

    for cell in ws[lemma_col_index[0]]:
        if (type(cell.value) != int) and (cell.value is not None) and (cell.value in ComVerblist):
            with open('TaggedData.txt', 'a') as file:
                file.write('{} \t {} \t {}\n'.format(cell.value, cell.row, os.path.basename(excel)))
            if not col_inserted:
                ws.insert_cols(cell.col_idx+1)
                merge_cells_in_new_column(ws, cell.col_idx+1, cell.col_idx)
                col_inserted = True
                ws.cell(row=1, column=cell.col_idx+1).value = 'ben'
                ws.cell(row=1, column=cell.col_idx + 1).font = copy(ws.cell(row=1, column=cell.col_idx).font)
                ws.cell(row=1, column=cell.col_idx + 1).fill = copy(ws.cell(row=1, column=cell.col_idx).fill)
                ws.cell(row=1, column=cell.col_idx + 1).border = copy(ws.cell(row=1, column=cell.col_idx).border)
                ws.cell(row=1, column=cell.col_idx + 1).alignment = copy(ws.cell(row=1, column=cell.col_idx).alignment)
            ws.cell(row=cell.row, column=cell.col_idx+1).value = 'verbal_process'

    wb.save(excel)

    print(os.path.basename(base_name), 'was tagged successfully')


# end function for optionality of input files
parser = argparse.ArgumentParser()
parser.add_argument('--table', nargs='*')
args = parser.parse_args()

headers = 'lemma\t cell \t file \n'
with open('TaggedData.txt', 'w') as file:
    file.write(headers)

# tag list of files in folder
if args.table:
    for f in args.table:
        process_file(f)
# tag all files in folder
else:
    excel_files = glob.glob(os.getcwd() + '/*.xlsx')


    for excel in excel_files:
        process_file(excel)

print('\n all selected files are successfully ComVerb-tagged now \n ฅ^•ﻌ•^ฅ')
